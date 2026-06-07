import re
import os
from datetime import datetime

def extract_table(text):
    """
    Extracts the markdown table from the given response text.
    [最小化优化]：放宽正则匹配条件，兼容中/英文表头，只匹配表格的分隔线结构。
    """
    # 优化后的正则：匹配类似 "|----|------|..." 这样的分隔线，并向下提取直到 "---" 结束。
    # (?<=\n) 确保匹配的是新的一行的开头
    table_pattern = r"(?<=\n)(\|[-]+\|[-]+\|.*)([\s\S]+?)(?=\n---)"
    
    match = re.search(table_pattern, text)

    if match:
        # group(1) 是那行 "|----|..." 分隔线，group(2) 是下面的数据行
        table_text = "| Dummy | Header | 行 |\n" + match.group(1) + match.group(2)
        return table_text.strip()
        
    return None


def clean_markdown(text):
    """
    Converts Markdown to plain text.
    - Removes **bold**, *italics*, and `<br>` tags.
    - Converts lists into readable format.
    """
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)  # Bold
    text = re.sub(r"\*(.*?)\*", r"\1", text)  # Italics
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)  # Replace <br>, <br/>, <BR /> with newlines
    text = re.sub(r"^\s*- ", "• ", text, flags=re.MULTILINE)  # Bullet points
    text = re.sub(r"^\s*(\d+)\.\s*", r"\1. ", text, flags=re.MULTILINE)  # Numbered lists
    text = "\n".join(line.strip() for line in text.splitlines())
    return text.strip()


def clean_tags(text):
    """
    Cleans only the tag column. Some models return tags like
    "{Sliding Window, Hash Map}", but braces may be meaningful in solution text,
    so we remove them only here.
    """
    text = clean_markdown(text).strip()
    if text.startswith("{") and text.endswith("}"):
        text = text[1:-1].strip()
    return text

def parse_markdown_table_rows(table_text):
    rows = []
    if not table_text:
        return rows
    
    for line in table_text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        columns = [clean_markdown(col.strip()) for col in line.split("|")[1:-1]]
        if not columns:
            continue

        first_col = columns[0].strip().lower()

        if first_col in {"dummy", "no.", "no", "题号"}:
            continue
        if all(re.fullmatch(r":?-+:?", col.strip()) for col in columns):
            continue
        if len(columns) >= 6:
            columns[3] = clean_tags(columns[3])
            rows.append(columns[:6])
    return rows

        


def parse_table_to_xlsx(table_text, output_xlsx="./data/leetcode_solutions.xlsx"):
    """
    Converts extracted markdown table into an Excel (.xlsx) file.
    ✅ Checks if the problem already exists, updates "Last Viewed" if found
    ✅ Appends new data if it does not exist
    ✅ Maintains formatting with Cambria font & text wrapping
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    headers = ["No.", "Name", "Last Viewed", "Tag", "Problem Pattern/Solution", "When to Use/Scale"]
    today_date = datetime.today().strftime('%Y-%m-%d')

    # Process new data using the shared markdown table parser.
    new_data = parse_markdown_table_rows(table_text)
    if not new_data:
        print("No valid table rows found.")
        return

    # Load existing Excel data (if file exists)
    if os.path.exists(output_xlsx):
        existing_df = pd.read_excel(output_xlsx, engine='openpyxl')

        # Convert "No." column to string for consistency
        existing_df["No."] = existing_df["No."].astype(str)

        for new_entry in new_data:
            problem_number, problem_title = new_entry[0], new_entry[1]

            # Check if the problem already exists
            match_idx = existing_df[(existing_df["No."] == problem_number) & (existing_df["Name"] == problem_title)].index

            if not match_idx.empty:
                # Update the "Last Viewed" column for the existing entry
                existing_df.loc[match_idx, "Last Viewed"] = today_date
                print(f"🔄 Updated Last Viewed for: LeetCode {problem_number}: {problem_title}")
            else:
                # Append new problem to DataFrame
                existing_df = pd.concat([existing_df, pd.DataFrame([new_entry], columns=headers)], ignore_index=True)
                print(f"➕ Added new entry: LeetCode {problem_number}: {problem_title}")

    else:
        existing_df = pd.DataFrame(new_data, columns=headers)

    # Save DataFrame to Excel
    existing_df.to_excel(output_xlsx, index=False, engine='openpyxl')

    # Load workbook and worksheet
    workbook = load_workbook(output_xlsx)
    worksheet = workbook.active

    # Define Cambria font styles
    cambria_font = Font(name="Cambria", size=11)
    cambria_bold_font = Font(name="Cambria", size=12, bold=True)  # For headers

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20 if column_title not in ["Problem Pattern/Solution",
                                                                                      "When to Use/Scale"] else 50

        # Apply formatting to headers
        header_cell = worksheet[column_letter + "1"]
        header_cell.font = cambria_bold_font  # Bold Cambria font for headers
        header_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Apply formatting to all data rows
        for row_num in range(2, len(existing_df) + 2):
            cell = worksheet[column_letter + str(row_num)]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = cambria_font

    # Save updated workbook
    workbook.save(output_xlsx)
    print(f"✅ Data updated in {output_xlsx} with 'Last Viewed' date.")


if __name__ == "__main__":
    # Example response containing markdown table
    response_text = """
    ### **LeetCode 1: Two Sum**  

    | No. | Name | Last Viewed | Tag | Problem Pattern/Solution | When to Use/Scale |
    |----|------|------------|-----|--------------------------|--------------------|
    | 1 | Two Sum | 2025-02-20 | Array, Hash Table | **Problem Pattern:** This problem involves finding two numbers in an array that add up to a specific target number. <br> **Solution Approach:** The solution uses a hash table to store numbers and their indices while iterating through the array. For each number, we check if the difference between the target and the current number is in the hash table. If it is, we return the current index and the index of the difference. | 1. Use this pattern when you need to find two numbers in an array that add up to a specific target. <br> 2. This pattern can be used when you need to solve problems that involve pairings in an array. <br> 3. This pattern is useful when a time-efficient solution is needed, as hash tables provide quick lookups. |
    """

    # Extract table and convert to Excel, appending if needed
    table_data = extract_table(response_text)
    if table_data:
        parse_table_to_xlsx(table_data)
    else:
        print("❌ No table found in response!")
